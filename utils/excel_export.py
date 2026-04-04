"""
Excel export utility for Moneytodo.
Generates a formatted report with summary and detailed breakdown.
"""
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def _border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color="4472C4"):
    return PatternFill("solid", fgColor=color)


def _alt_fill(color="DCE6F1"):
    return PatternFill("solid", fgColor=color)


def export_monthly_report(month: str, expenses: list, summary: dict, category_summary: list, output_dir: str = None) -> str:
    """
    Export a monthly report to Excel.

    :param month: YYYY-MM string
    :param expenses: list of expense dicts (all rows for the month)
    :param summary: dict with keys 'income' and 'expense'
    :param category_summary: list of per-category summary dicts
    :param output_dir: directory to save the file (default: current directory)
    :return: full path of the saved file
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    if output_dir is None:
        output_dir = os.path.expanduser("~")

    try:
        month_label = datetime.strptime(month, "%Y-%m").strftime("%m/%Y")
    except ValueError:
        raise ValueError(f"Invalid month format. Expected YYYY-MM, got: {month}")
    filename = f"BaoCao_ThiChi_{month.replace('-', '_')}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tóm tắt"
    _write_summary_sheet(ws1, month_label, summary, category_summary)

    # ── Sheet 2: Transactions ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chi Tiết Giao Dịch")
    _write_transactions_sheet(ws2, month_label, expenses)

    wb.save(filepath)
    return filepath


def _write_summary_sheet(ws, month_label, summary, category_summary):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO THU CHI THÁNG {month_label}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _header_fill("1F497D")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3
    # Summary box
    labels = [("Tổng Thu", summary.get("income", 0)), ("Tổng Chi", summary.get("expense", 0))]
    balance = summary.get("income", 0) - summary.get("expense", 0)
    labels.append(("Số Dư", balance))

    colors = ["70AD47", "FF0000", "4472C4"]
    for i, (label, val) in enumerate(labels):
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_l.font = Font(bold=True, color="FFFFFF")
        cell_l.fill = PatternFill("solid", fgColor=colors[i])
        cell_l.border = _border()
        cell_l.alignment = Alignment(horizontal="center")

        cell_v = ws.cell(row=row, column=2, value=val)
        cell_v.number_format = '#,##0 "₫"'
        cell_v.border = _border()
        cell_v.alignment = Alignment(horizontal="right")
        row += 1

    row += 1
    # Category breakdown header
    headers = ["Danh Mục", "Danh Mục Con", "Loại", "Tổng Tiền"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, item in enumerate(category_summary):
        fill = _alt_fill() if i % 2 == 0 else PatternFill()
        vals = [
            item.get("category") or "—",
            item.get("subcategory") or "—",
            "Thu" if item.get("type") == "income" else "Chi",
            item.get("total", 0),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border()
            if col == 4:
                cell.number_format = '#,##0 "₫"'
                cell.alignment = Alignment(horizontal="right")
            elif col == 3:
                cell.font = Font(
                    color="008000" if item.get("type") == "income" else "FF0000"
                )
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill
        row += 1


def _write_transactions_sheet(ws, month_label, expenses):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 30

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"CHI TIẾT GIAO DỊCH THÁNG {month_label}"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = _header_fill("1F497D")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["STT", "Ngày", "Loại", "Danh Mục", "Danh Mục Con", "Số Tiền", "Diễn Giải"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")

    for i, exp in enumerate(expenses, 1):
        row = i + 2
        fill = _alt_fill() if i % 2 == 0 else PatternFill()
        type_label = "Thu" if exp.get("type") == "income" else "Chi"
        row_data = [
            i,
            exp.get("expense_date", ""),
            type_label,
            exp.get("category_name") or "—",
            exp.get("subcategory_name") or "—",
            exp.get("amount", 0),
            exp.get("description") or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border()
            if col == 6:
                cell.number_format = '#,##0 "₫"'
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(
                    color="008000" if exp.get("type") == "income" else "FF0000"
                )
            elif col == 3:
                cell.font = Font(
                    color="008000" if exp.get("type") == "income" else "FF0000"
                )
            if i % 2 == 0:
                cell.fill = fill
